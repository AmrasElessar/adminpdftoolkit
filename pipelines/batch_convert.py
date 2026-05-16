"""Batch-Excel merge pipeline.

Was the 190-line ``_batch_convert_worker`` god function in ``app.py``;
hoisted to its own module during S5.2 along with the per-token JSON
load / save / pipeline helpers it depends on.

Public surface:

* :func:`batch_convert_worker` — the actual thread entrypoint.
* :func:`load_job` / :func:`save_view` — read / mutate the per-token JSON.
* :func:`apply_pipeline` — re-derive the displayed records from the
  originals when dedupe / filter state changes.
* :func:`load_distribution` — read the distribution JSON.
* :func:`write_merged_excel` — single Excel writer shared by the merge
  endpoint, the view-save flow, and the per-team distribution download.

The router and the worker import from here; ``app.py`` is no longer
involved.
"""

from __future__ import annotations

import json
import os
import shutil
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import core
from core import logger, sanitize_error
from settings import settings
from state import batch_store

_CALL_LOG_COL_WIDTHS = {
    "Sıra": 7,
    "Kaynak PDF": 28,
    "Kayıt No": 10,
    "Müşteri": 28,
    "Telefon": 18,
    "Durum": 12,
    "Tarih": 18,
    "Süre": 10,
    "Ağrı / romatizma": 22,
    "Termal / kaplıca": 22,
    "Yaş": 14,
    "Medeni durum": 16,
    "Meslek": 22,
    "İkamet ili": 16,
    "AI Özeti (Ham)": 50,
}


def write_merged_excel(
    records: list[dict],
    source_files: list[str],
    out_path: Path,
    *,
    schema: list[str] | None = None,
    sheet_title: str = "Birleşik Kayıtlar",
    unsafe: bool = False,
) -> None:
    """Write the merged-records workbook used by /batch-download and the
    per-team distribution ZIP.

    ``schema`` is the list of headers to write *between* ``Sıra/Kaynak PDF``
    and (optionally) ``AI Özeti (Ham)``. When omitted, falls back to the
    live ``core.TARGET_SCHEMA`` + the trailing ``AI Özeti (Ham)`` column
    (call-log layout). For ``other_table`` groups, the caller passes the
    group's own headers and we skip the AI-summary trailing column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    if schema is None:
        target_schema = list(getattr(core, "TARGET_SCHEMA", []))
        headers = ["Sıra", "Kaynak PDF", "Kayıt No", *target_schema, "AI Özeti (Ham)"]
    else:
        headers = ["Sıra", "Kaynak PDF", *schema]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row_idx, (rec, src) in enumerate(zip(records, source_files, strict=False), 2):
        orig_sira = rec.get("Sıra", row_idx - 1)
        for col_idx, h in enumerate(headers, 1):
            if h == "Sıra":
                val = orig_sira
            elif h == "Kaynak PDF":
                val = src
            else:
                val = rec.get(h, "")
            if isinstance(val, str) and "\n" in val:
                val = val.replace("\r\n", " ").replace("\n", " ").strip()
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=False)

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _CALL_LOG_COL_WIDTHS.get(h, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if unsafe:
        from pdf_converter import _add_excel_unsafe_sheet

        _add_excel_unsafe_sheet(wb)
    wb.save(out_path)


def _dedup_key(rec: dict, match_columns: list[str]) -> str | None:
    """Build a composite key for dedup. For the call-log default
    (``["Telefon"]``) we still pass through ``normalize_phone`` so that
    ``+90 555 …`` and ``05555…`` collapse. For user-picked columns
    (other_table groups) we lowercase-strip each field and join with a
    delimiter that can't appear in cell text. ``None`` means "skip this
    row" (no key info → don't dedup it away)."""
    if match_columns == ["Telefon"]:
        return core.normalize_phone(rec.get("Telefon")) or None
    parts = []
    has_any = False
    for col in match_columns:
        v = str(rec.get(col, "")).strip().lower()
        if v:
            has_any = True
        parts.append(v)
    if not has_any:
        return None
    return "\x1f".join(parts)


def apply_pipeline(
    original_records: list[dict],
    original_sources: list[str],
    state: dict,
) -> tuple[list[dict], list[str]]:
    """Re-derive (records, sources) from the originals based on ``state``
    (dedupe + filters). Original lists are never mutated."""
    records = [dict(r) for r in original_records]
    sources = list(original_sources)

    if state.get("deduplicated"):
        match_columns = state.get("match_columns") or ["Telefon"]
        seen: set[str] = set()
        nrec, nsrc = [], []
        for rec, src in zip(records, sources, strict=False):
            k = _dedup_key(rec, match_columns)
            if k and k in seen:
                continue
            if k:
                seen.add(k)
            nrec.append(rec)
            nsrc.append(src)
        records, sources = nrec, nsrc

    filters = state.get("filters") or {}
    if filters:
        nrec, nsrc = [], []
        for rec, src in zip(records, sources, strict=False):
            ok = True
            for col, allowed in filters.items():
                if not allowed:
                    continue
                val = str(rec.get(col, "")).strip()
                if val == "":
                    if "(boş)" not in allowed:
                        ok = False
                        break
                else:
                    if val not in allowed:
                        ok = False
                        break
            if ok:
                nrec.append(rec)
                nsrc.append(src)
        records, sources = nrec, nsrc

    for i, rec in enumerate(records, start=1):
        rec["Sıra"] = i

    return records, sources


def load_job(token: str) -> dict[str, Any]:
    """Load ``data.json`` for a batch-merge token, migrating old shape.

    Modern ``data.json`` (written by ``batch_convert_worker``) carries
    a ``group_kind`` (``"call_log"`` or ``"other_table"``) and an explicit
    ``group_headers`` list — these drive the dedup default + Excel writer
    schema. Older shapes get migrated to the call-log defaults.
    """
    core.check_token(token)
    job_dir = core.make_job_dir("jobs", token)
    data_path = job_dir / "data.json"
    if not data_path.exists():
        raise HTTPException(404, "İş bulunamadı veya süresi dolmuş.")
    with data_path.open("r", encoding="utf-8") as fp:
        data: dict[str, Any] = json.load(fp)
    migrated = False
    if "original_records" not in data:
        data["original_records"] = list(data.get("records", []))
        data["original_sources"] = list(data.get("source_files", []))
        migrated = True
    if "state" not in data:
        data["state"] = {"deduplicated": False, "filters": {}}
        migrated = True
    if "group_kind" not in data:
        data["group_kind"] = "call_log"
        migrated = True
    if "group_headers" not in data:
        # call-log default: TARGET_SCHEMA bracketed by Kayıt No / AI Özeti (Ham)
        target_schema = list(getattr(core, "TARGET_SCHEMA", []))
        data["group_headers"] = ["Kayıt No", *target_schema, "AI Özeti (Ham)"]
        migrated = True
    if "group_label" not in data:
        data["group_label"] = "Çağrı kayıtları" if data["group_kind"] == "call_log" else "Tablo"
        migrated = True
    # State sub-fields
    st = data["state"]
    if "match_columns" not in st:
        st["match_columns"] = ["Telefon"] if data["group_kind"] == "call_log" else []
        migrated = True
    if migrated:
        with data_path.open("w", encoding="utf-8") as fp:
            json.dump(data, fp, ensure_ascii=False)
    return data


def save_view(token: str, state: dict) -> dict:
    """Run the pipeline for ``state``, rewrite the Excel + ``data.json``,
    and drop any stale per-team distribution file. Returns a small summary
    dict for the caller to forward to the frontend."""
    job_dir = core.make_job_dir("jobs", token)
    data = load_job(token)
    records, sources = apply_pipeline(data["original_records"], data["original_sources"], state)

    for old in job_dir.glob("birlesik_*.xlsx"):
        old.unlink(missing_ok=True)
    out_name = f"birlesik_{len(records)}_kayit.xlsx"
    out = job_dir / out_name
    schema = data["group_headers"] if data["group_kind"] == "other_table" else None
    write_merged_excel(records, sources, out, schema=schema)

    data["records"] = records
    data["source_files"] = sources
    data["filename"] = out_name
    data["state"] = state
    with (job_dir / "data.json").open("w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False)

    dist = job_dir / "distribution.json"
    if dist.exists():
        dist.unlink(missing_ok=True)

    return {
        "record_count": len(records),
        "filename": out_name,
        "state": state,
    }


def load_distribution(token: str) -> dict[str, Any]:
    core.check_token(token)
    job_dir = core.make_job_dir("jobs", token)
    path = job_dir / "distribution.json"
    if not path.exists():
        raise HTTPException(404, "Dağıtım bulunamadı — önce 'Dağıt' yapın.")
    with path.open("r", encoding="utf-8") as fp:
        loaded: dict[str, Any] = json.load(fp)
    return loaded


def _set_file_progress(
    token: str,
    fp_idx_by_name: dict[str, int],
    fname: str,
    **fields: Any,
) -> None:
    """Mutate one row of ``files_progress`` under the batch lock.

    ``files_progress`` is a list inside the job dict; it's safe to mutate
    in-place because every reader copies the dict (and the list with it)
    while holding ``batch_store.lock``.
    """
    with batch_store.lock:
        job = batch_store.jobs.get(token)
        if job is None:
            return
        fp_list = job.get("files_progress") or []
        idx = fp_idx_by_name.get(fname)
        if idx is None or idx >= len(fp_list):
            return
        for k, v in fields.items():
            fp_list[idx][k] = v


def batch_convert_worker(
    token: str,
    files_data: list[tuple[str, Path]],
    mappings_obj: dict,
    skip_list: list[str],
    job_token: str,
    file_count: int,
    skip_safety: bool = False,
    *,
    group_kind: str = "call_log",
    group_headers: list[str] | None = None,
    group_label: str | None = None,
) -> None:
    """Excel-merge worker. Parses every (non-skipped) PDF in parallel via a
    ``ProcessPoolExecutor`` when ``settings.parallel_batch_workers`` allows
    (auto = ``min(cpu_count, 4)``); falls back to serial parsing for small
    batches or pool-unavailable environments. Output ordering is
    deterministic — completion order is stored back at the original index
    so the merged Excel always matches the user-supplied PDF order.

    ``group_kind`` is either ``"call_log"`` or ``"other_table"``. For
    other_table the caller passes ``group_headers`` (the common column
    names detected by ``/batch-analyze``); records are keyed by those
    headers instead of being squashed into the call-log schema.
    """
    job_dir = core.make_job_dir("jobs", job_token)
    merged: list[dict[str, Any]] = []
    source_index: list[str] = []
    warnings: list[str] = []
    target_schema = list(getattr(core, "TARGET_SCHEMA", []))
    if group_kind == "other_table":
        if not group_headers:
            batch_store.update(
                token,
                error="other_table grubu için sütun başlıkları (group_headers) gerekli.",
                done=True,
            )
            return
        effective_headers = list(group_headers)
        effective_label = group_label or "Tablo"
    else:
        effective_headers = ["Kayıt No", *target_schema, "AI Özeti (Ham)"]
        effective_label = group_label or "Çağrı kayıtları"
    try:
        # Phase 0: per-file safety scan (cancellable). Pre-conversion gate
        # used to be in the request handler; doing it here keeps the user
        # informed via SSE and lets them press "Atla" to skip remaining.
        unsafe_any = False
        if not skip_safety:
            from pipelines.safety import scan_files_with_progress

            scan_targets = [(fn, p) for fn, p in files_data if fn not in skip_list]
            if scan_targets:
                ok, err = scan_files_with_progress(batch_store, token, scan_targets)
                if not ok:
                    batch_store.update(token, error=err or "Güvensiz PDF reddedildi.", done=True)
                    return
                snap = batch_store.snapshot(token) or {}
                unsafe_any = bool(snap.get("unsafe_files"))

        files_progress = [
            {
                "name": fn,
                "status": "skipped" if fn in skip_list else "pending",
                "record_count": 0,
                "error": None,
            }
            for fn, _ in files_data
        ]
        batch_store.update(token, phase="processing", files_progress=files_progress)
        fp_idx_by_name: dict[str, int] = {str(fp["name"]): i for i, fp in enumerate(files_progress)}

        parse_args: list[tuple] = []
        for filename, pdf_path in files_data:
            if filename in skip_list:
                continue
            # mode tells the parser how to shape records. For other_table
            # groups we pass the group's headers as the schema (the parser
            # extracts the table verbatim and keys rows by those headers).
            mode = group_kind
            schema_for_parse = effective_headers if group_kind == "other_table" else target_schema
            parse_args.append(
                (filename, str(pdf_path), mappings_obj.get(filename), schema_for_parse, mode)
            )

        configured = settings.parallel_batch_workers
        workers = min(max(1, os.cpu_count() or 1), 4) if configured <= 0 else max(1, configured)

        use_pool = workers > 1 and len(parse_args) > 3
        results_in_order: list[dict[str, Any] | None] = [None] * len(parse_args)

        for a in parse_args:
            _set_file_progress(token, fp_idx_by_name, a[0], status="processing")

        if use_pool:
            try:
                with ProcessPoolExecutor(max_workers=workers) as executor:
                    futures = {
                        executor.submit(core.parse_pdf_for_batch, a): idx
                        for idx, a in enumerate(parse_args)
                    }
                    for completed, fut in enumerate(as_completed(futures), start=1):
                        idx = futures[fut]
                        try:
                            results_in_order[idx] = fut.result()
                        except Exception as e:
                            results_in_order[idx] = {
                                "filename": parse_args[idx][0],
                                "records": [],
                                "warning": f"{parse_args[idx][0]} atlandı: {sanitize_error(e)}",
                            }
                        result = results_in_order[idx]
                        assert result is not None  # just assigned above
                        rec_count = len(result.get("records", []))
                        warning = result.get("warning")
                        _set_file_progress(
                            token,
                            fp_idx_by_name,
                            result["filename"],
                            status="error" if warning and not rec_count else "done",
                            record_count=rec_count,
                            error=warning,
                        )
                        batch_store.update(
                            token,
                            current=completed,
                            last_file=result["filename"],
                        )
            except Exception as pool_err:
                logger.warning(
                    "ProcessPoolExecutor unusable, falling back to serial: %s",
                    sanitize_error(pool_err),
                )
                use_pool = False

        if not use_pool:
            for idx, a in enumerate(parse_args):
                try:
                    results_in_order[idx] = core.parse_pdf_for_batch(a)
                except Exception as e:
                    results_in_order[idx] = {
                        "filename": a[0],
                        "records": [],
                        "warning": f"{a[0]} atlandı: {sanitize_error(e)}",
                    }
                result = results_in_order[idx]
                assert result is not None  # just assigned above
                rec_count = len(result.get("records", []))
                warning = result.get("warning")
                _set_file_progress(
                    token,
                    fp_idx_by_name,
                    result["filename"],
                    status="error" if warning and not rec_count else "done",
                    record_count=rec_count,
                    error=warning,
                )
                batch_store.update(
                    token,
                    current=idx + 1,
                    last_file=result["filename"],
                )

        for r in results_in_order:
            if r is None:
                continue
            if r.get("warning"):
                warnings.append(r["warning"])
            for rec in r.get("records", []):
                merged.append(rec)
                source_index.append(r["filename"])

        if not merged:
            raise RuntimeError(
                "Birleştirilecek veri bulunamadı. Uyumsuz PDF'ler için sütunları eşleyin."
            )

        batch_store.update(token, phase="writing")

        for i, rec in enumerate(merged, start=1):
            rec["Sıra"] = i

        out_name = (
            f"birlesik_{len(merged)}_kayit_GUVENSIZ.xlsx"
            if unsafe_any
            else f"birlesik_{len(merged)}_kayit.xlsx"
        )
        out = job_dir / out_name
        schema_for_write = effective_headers if group_kind == "other_table" else None
        write_merged_excel(
            merged,
            source_index,
            out,
            schema=schema_for_write,
            unsafe=unsafe_any,
        )

        default_match_columns = ["Telefon"] if group_kind == "call_log" else []
        data_path = job_dir / "data.json"
        with data_path.open("w", encoding="utf-8") as fp:
            json.dump(
                {
                    "records": merged,
                    "source_files": source_index,
                    "filename": out_name,
                    "original_records": [dict(r) for r in merged],
                    "original_sources": list(source_index),
                    "state": {
                        "deduplicated": False,
                        "filters": {},
                        "match_columns": default_match_columns,
                    },
                    "group_kind": group_kind,
                    "group_headers": effective_headers,
                    "group_label": effective_label,
                },
                fp,
                ensure_ascii=False,
            )

        for p in job_dir.glob("in_*.pdf"):
            p.unlink(missing_ok=True)

        result = {
            "token": job_token,
            "record_count": len(merged),
            "source_count": file_count,
            "skipped_count": len(skip_list),
            "warnings": warnings,
            "filename": out_name,
            "group_kind": group_kind,
            "group_headers": effective_headers,
            "group_label": effective_label,
        }
        batch_store.update(token, result=result, phase="done", done=True)
    except Exception as e:
        logger.exception("batch-convert worker failed (token=%s)", token)
        batch_store.update(token, error=sanitize_error(e), done=True)
        shutil.rmtree(job_dir, ignore_errors=True)
    finally:
        final_snap = batch_store.snapshot(token)
        if final_snap:
            core.persist_job_state("batch", token, final_snap)
