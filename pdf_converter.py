"""
PDF Dönüştürücü - Şirket İçi Araç
---------------------------------
Ana dizindeki PDF'leri Excel / Word / JPG formatına çevirir.
Tamamen offline çalışır, limit yoktur, dosyalar dışarı çıkmaz.

S2 sonrası: PDF kind detection (call-log / scanned) ve call-log parsing
``parsers/`` paketine taşındı. Bu modül artık Excel/Word/JPG yazıcılarını
ve OCR'ı barındırıyor; classify/parse fonksiyonları aşağıda parsers'a
delegate eden shim olarak korunuyor (geriye uyum).

Heavy imports (``openpyxl``, ``pdf2docx``) are deferred into the functions
that use them so the spawn-mode batch worker (see ``core.batch
.parse_pdf_for_batch``) doesn't drag the full 50+ MB chain on every spawn —
the worker only needs ``is_call_log_pdf`` / ``parse_call_log``.
"""

from __future__ import annotations

import sys
from pathlib import Path

import fitz  # PyMuPDF — used in type hints + by parsers; cheap-ish, leave at top.

from parsers.call_log_360 import (
    CALL_LOG_QUESTIONS,
    CallLog360Parser,
)
from parsers.scanned import ScannedParser


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ----------------------------------------------------------------------------
# Backward-compat shims — eski importlar (`from pdf_converter import ...`)
# kırılmasın diye parsers paketine delegate ediyoruz.
# ----------------------------------------------------------------------------

_call_log_parser = CallLog360Parser()
_scanned_parser = ScannedParser()


def is_call_log_pdf(doc: fitz.Document) -> bool:
    return _call_log_parser.is_match(doc)


def is_scanned_pdf(doc: fitz.Document) -> bool:
    return _scanned_parser.is_match(doc)


def parse_call_log(doc: fitz.Document) -> list[dict]:
    return _call_log_parser.parse_records(doc)


# ----------------------------------------------------------------------------
# Excel yazımı
# ----------------------------------------------------------------------------

def write_call_log_excel(records: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Çağrı Kayıtları"

    headers = ["Sıra", "Kayıt No", "Müşteri", "Telefon", "Durum", "Tarih", "Süre"] + CALL_LOG_QUESTIONS + ["AI Özeti (Ham)"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row_idx, rec in enumerate(records, 2):
        sira = row_idx - 1  # 1'den başlayıp artan
        for col_idx, h in enumerate(headers, 1):
            if h == "Sıra":
                val = sira
            elif h == "Kayıt No":
                val = rec.get("#", "")
            else:
                val = rec.get(h, "")
            # AI Özeti (Ham) alanındaki satır sonlarını boşlukla değiştir ki
            # hücre tek satırda görünsün — metin kaydır aktif olmasın.
            if isinstance(val, str) and "\n" in val:
                val = val.replace("\r\n", " ").replace("\n", " ").strip()
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=False)

    # Sütun genişlikleri
    widths = {
        "Sıra": 7, "Kayıt No": 10, "Müşteri": 28, "Telefon": 18, "Durum": 12, "Tarih": 18, "Süre": 10,
        "Ağrı / romatizma": 22, "Termal / kaplıca": 22, "Yaş": 14,
        "Medeni durum": 16, "Meslek": 22, "İkamet ili": 16, "AI Özeti (Ham)": 50,
    }
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


def write_generic_excel(doc: fitz.Document, out_path: Path) -> None:
    """Çağrı kaydı olmayan PDF'ler için — her sayfayı ayrı sheet'e metin olarak yazar."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for i, page in enumerate(doc, 1):
        ws = wb.create_sheet(title=f"Sayfa {i}"[:31])
        lines = page.get_text().splitlines()
        for row_idx, ln in enumerate(lines, 1):
            ws.cell(row=row_idx, column=1, value=ln)
        ws.column_dimensions["A"].width = 120
    wb.save(out_path)


# ----------------------------------------------------------------------------
# Word ve JPG dönüşümleri
# ----------------------------------------------------------------------------

def convert_to_word(pdf_path: Path, out_path: Path) -> None:
    from pdf2docx import Converter as Pdf2DocxConverter

    cv = Pdf2DocxConverter(str(pdf_path))
    try:
        cv.convert(str(out_path), start=0, end=None)
    finally:
        cv.close()


# ----------------------------------------------------------------------------
# OCR — taranmış/görsel PDF'ler için
# ----------------------------------------------------------------------------

_ocr_reader = None


def get_ocr_reader():
    """EasyOCR okuyucusunu lazy olarak başlatır.

    Portable kurulumda: proje klasöründe `_EasyOCR_models/` varsa o kullanılır,
    indirme kapatılır. Normal kurulumda model ~/.EasyOCR/ altına indirilir.
    """
    global _ocr_reader
    if _ocr_reader is None:
        import easyocr  # type: ignore

        kwargs: dict[str, object] = {"gpu": False, "verbose": False}
        local_models = Path(__file__).parent / "_EasyOCR_models"
        if local_models.exists():
            kwargs["model_storage_directory"] = str(local_models)
            user_dir = local_models / "user_network"
            user_dir.mkdir(exist_ok=True)
            kwargs["user_network_directory"] = str(user_dir)
            kwargs["download_enabled"] = False  # offline çalış
        _ocr_reader = easyocr.Reader(["tr", "en"], **kwargs)
    return _ocr_reader


def ocr_pdf_pages(pdf_path: Path, dpi: int = 200) -> list[str]:
    """Her sayfayı OCR'dan geçirip metin döndürür."""
    reader = get_ocr_reader()
    zoom = dpi / 72
    mat = fitz.Matrix(zoom, zoom)
    doc = fitz.open(str(pdf_path))
    out: list[str] = []
    try:
        for page in doc:
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("png")
            pix = None  # release Pixmap eagerly between pages
            lines = reader.readtext(img_bytes, detail=0, paragraph=True) or []
            out.append("\n".join(lines))
    finally:
        doc.close()
    return out


def convert_to_jpg(pdf_path: Path, out_dir: Path, dpi: int = 200) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    zoom = dpi / 72
    mat = fitz.Matrix(zoom, zoom)
    doc = fitz.open(str(pdf_path))
    written: list[Path] = []
    try:
        for i, page in enumerate(doc, 1):
            pix = page.get_pixmap(matrix=mat, alpha=False)
            out = out_dir / f"sayfa_{i:03d}.jpg"
            pix.save(str(out), jpg_quality=90)
            pix = None  # release Pixmap eagerly between pages
            written.append(out)
    finally:
        doc.close()
    return written


# ----------------------------------------------------------------------------
# CLI (yardımcı — pdf_converter.py doğrudan çalıştırılırsa)
# Web sunucusu app.py üzerinden çalışır; bu modülün GUI kısmı kaldırıldı.
# ----------------------------------------------------------------------------


def main() -> None:
    if len(sys.argv) >= 3:
        # CLI: python pdf_converter.py <pdf> <excel|word|jpg>
        pdf = Path(sys.argv[1])
        fmt = sys.argv[2].lower()
        stem = pdf.stem
        if fmt == "excel":
            doc = fitz.open(str(pdf))
            try:
                out = OUTPUT_DIR / f"{stem}.xlsx"
                if is_call_log_pdf(doc):
                    recs = parse_call_log(doc)
                    write_call_log_excel(recs, out)
                    print(f"[OK] {len(recs)} kayit -> {out}")
                else:
                    write_generic_excel(doc, out)
                    print(f"[OK] Sayfa sayfa -> {out}")
            finally:
                doc.close()
        elif fmt == "word":
            out = OUTPUT_DIR / f"{stem}.docx"
            convert_to_word(pdf, out)
            print(f"[OK] {out}")
        elif fmt == "jpg":
            out_dir = OUTPUT_DIR / stem
            files = convert_to_jpg(pdf, out_dir)
            print(f"[OK] {len(files)} sayfa -> {out_dir}")
        else:
            print(f"Bilinmeyen format: {fmt}")
            sys.exit(1)
        return

    print("Bu modül web sunucusu olarak app.py üzerinden çalıştırılır.")
    print("Doğrudan kullanım: python pdf_converter.py <pdf> <excel|word|jpg>")


if __name__ == "__main__":
    main()
