"""Unit tests for the batch-merge pipeline helpers.

Lifted from ``app.py`` during S5.2 into ``pipelines/batch_convert.py``
along with the actual ``batch_convert_worker``. ``apply_pipeline`` is
already covered by ``test_pipeline.py``; this file pins the rest:

* :func:`load_job` — reads ``data.json``, transparently migrates the
  pre-S5 shape to the current (``original_*`` / ``state``) shape.
* :func:`save_view` — applies dedupe + filters, writes Excel +
  ``data.json``, drops a stale distribution file.
* :func:`load_distribution` — 404s when there's no distribution.
* :func:`write_merged_excel` — produces a real workbook with the
  expected headers and freeze panes.

Each test gets its own job-dir under ``_work/jobs/<random>`` and tears
it down at the end, so they don't leak state across runs.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

import core
from pipelines.batch_convert import (
    load_distribution,
    load_job,
    save_view,
    write_merged_excel,
)


@pytest.fixture
def job_token() -> str:
    """Brand new uuid4().hex token — same shape the routers use."""
    return uuid4().hex


@pytest.fixture
def job_dir(job_token: str):
    """Create the per-job directory and tear it down on exit."""
    d = core.make_job_dir("jobs", job_token)
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _seed_data_json(job_dir: Path, payload: dict) -> None:
    (job_dir / "data.json").write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


# ---------------------------------------------------------------------------
# load_job — current shape, missing file, legacy migration
# ---------------------------------------------------------------------------
def test_load_job_returns_current_shape(job_dir: Path, job_token: str) -> None:
    _seed_data_json(
        job_dir,
        {
            "records": [{"Telefon": "1"}, {"Telefon": "2"}],
            "source_files": ["a.pdf", "b.pdf"],
            "filename": "out.xlsx",
            "original_records": [{"Telefon": "1"}, {"Telefon": "2"}],
            "original_sources": ["a.pdf", "b.pdf"],
            "state": {"deduplicated": False, "filters": {}},
        },
    )
    data = load_job(job_token)
    assert len(data["records"]) == 2
    assert data["filename"] == "out.xlsx"
    assert data["state"]["deduplicated"] is False


def test_load_job_404_when_missing(job_token: str) -> None:
    """No data.json on disk → HTTPException(404). Routers rely on this
    to translate "expired/unknown token" into a clean error response."""
    with pytest.raises(HTTPException) as exc:
        load_job(job_token)
    assert exc.value.status_code == 404


def test_load_job_migrates_legacy_shape(job_dir: Path, job_token: str) -> None:
    """Pre-S5 ``data.json`` had only ``records`` / ``source_files`` —
    ``load_job`` upgrades it in place by seeding ``original_*`` and a
    default empty pipeline state. The on-disk file should also be rewritten
    so subsequent loads skip the migration path."""
    _seed_data_json(
        job_dir,
        {
            "records": [{"Telefon": "555"}],
            "source_files": ["x.pdf"],
            "filename": "old.xlsx",
        },
    )
    data = load_job(job_token)
    assert data["original_records"] == [{"Telefon": "555"}]
    assert data["original_sources"] == ["x.pdf"]
    assert data["state"] == {"deduplicated": False, "filters": {}}

    # Re-read from disk to confirm migration was persisted.
    raw = json.loads((job_dir / "data.json").read_text(encoding="utf-8"))
    assert "original_records" in raw
    assert raw["state"]["deduplicated"] is False


def test_load_job_seeds_state_when_missing(job_dir: Path, job_token: str) -> None:
    """Variant of the legacy migration: ``original_*`` already present
    but ``state`` was forgotten. Load must add the empty default."""
    _seed_data_json(
        job_dir,
        {
            "records": [{}],
            "source_files": ["x.pdf"],
            "filename": "y.xlsx",
            "original_records": [{}],
            "original_sources": ["x.pdf"],
            # state intentionally absent
        },
    )
    data = load_job(job_token)
    assert data["state"] == {"deduplicated": False, "filters": {}}


# ---------------------------------------------------------------------------
# save_view — applies pipeline, rewrites Excel, drops distribution
# ---------------------------------------------------------------------------
def test_save_view_dedupe_rewrites_excel_and_data(job_dir: Path, job_token: str) -> None:
    _seed_data_json(
        job_dir,
        {
            "records": [
                {"Sıra": 1, "Telefon": "5551112233", "Müşteri": "Ali"},
                {"Sıra": 2, "Telefon": "5551112233", "Müşteri": "Ali (dup)"},
                {"Sıra": 3, "Telefon": "5552223344", "Müşteri": "Veli"},
            ],
            "source_files": ["a.pdf", "a.pdf", "b.pdf"],
            "filename": "birlesik_3_kayit.xlsx",
            "original_records": [
                {"Sıra": 1, "Telefon": "5551112233", "Müşteri": "Ali"},
                {"Sıra": 2, "Telefon": "5551112233", "Müşteri": "Ali (dup)"},
                {"Sıra": 3, "Telefon": "5552223344", "Müşteri": "Veli"},
            ],
            "original_sources": ["a.pdf", "a.pdf", "b.pdf"],
            "state": {"deduplicated": False, "filters": {}},
        },
    )
    # Pre-create the previous Excel so we can confirm it gets cleaned up.
    (job_dir / "birlesik_3_kayit.xlsx").write_bytes(b"old")

    result = save_view(job_token, {"deduplicated": True, "filters": {}})

    assert result["record_count"] == 2  # one duplicate dropped
    assert result["filename"] == "birlesik_2_kayit.xlsx"
    assert (job_dir / "birlesik_2_kayit.xlsx").exists()
    # Old Excel must be gone — save_view sweeps any prior birlesik_*.xlsx.
    assert not (job_dir / "birlesik_3_kayit.xlsx").exists()

    # data.json should now reflect the new view.
    raw = json.loads((job_dir / "data.json").read_text(encoding="utf-8"))
    assert raw["filename"] == "birlesik_2_kayit.xlsx"
    assert raw["state"]["deduplicated"] is True
    # Sıra is renumbered after the pipeline runs.
    assert [r["Sıra"] for r in raw["records"]] == [1, 2]


def test_save_view_drops_stale_distribution(job_dir: Path, job_token: str) -> None:
    """Editing the merged view invalidates any prior distribution snapshot;
    ``save_view`` deletes ``distribution.json`` so the UI doesn't keep
    showing assignments built on records that have since been filtered."""
    _seed_data_json(
        job_dir,
        {
            "records": [{"Telefon": "1"}],
            "source_files": ["x.pdf"],
            "filename": "out.xlsx",
            "original_records": [{"Telefon": "1"}],
            "original_sources": ["x.pdf"],
            "state": {"deduplicated": False, "filters": {}},
        },
    )
    (job_dir / "distribution.json").write_text("{}", encoding="utf-8")

    save_view(job_token, {"deduplicated": False, "filters": {}})
    assert not (job_dir / "distribution.json").exists()


# ---------------------------------------------------------------------------
# load_distribution — 404 + roundtrip
# ---------------------------------------------------------------------------
def test_load_distribution_404_when_missing(job_dir: Path, job_token: str) -> None:
    # job_dir exists but distribution.json doesn't
    with pytest.raises(HTTPException) as exc:
        load_distribution(job_token)
    assert exc.value.status_code == 404


def test_load_distribution_returns_payload(job_dir: Path, job_token: str) -> None:
    payload = {
        "strategy": "sequential",
        "teams": ["A", "B"],
        "assignments": [
            {"name": "A", "records": [{}], "sources": ["a.pdf"]},
            {"name": "B", "records": [], "sources": []},
        ],
    }
    (job_dir / "distribution.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )
    out = load_distribution(job_token)
    assert out["strategy"] == "sequential"
    assert out["assignments"][0]["name"] == "A"


# ---------------------------------------------------------------------------
# write_merged_excel — workbook structure
# ---------------------------------------------------------------------------
def test_write_merged_excel_writes_expected_header_row(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Workbook starts with: Sıra, Kaynak PDF, Kayıt No, <TARGET_SCHEMA…>,
    AI Özeti (Ham). app.py mirrors the live schema onto core.TARGET_SCHEMA
    at import time; pin a deterministic copy here to keep this test
    self-contained."""
    monkeypatch.setattr(core, "TARGET_SCHEMA", ["Müşteri", "Telefon", "Durum"])

    out = tmp_path / "merged.xlsx"
    write_merged_excel(
        records=[
            {"Sıra": 1, "Müşteri": "Ali", "Telefon": "1", "Durum": "ended"},
            {"Sıra": 2, "Müşteri": "Veli", "Telefon": "2", "Durum": "missed"},
        ],
        source_files=["a.pdf", "b.pdf"],
        out_path=out,
    )
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == [
        "Sıra",
        "Kaynak PDF",
        "Kayıt No",
        "Müşteri",
        "Telefon",
        "Durum",
        "AI Özeti (Ham)",
    ]
    # Two data rows
    assert ws.max_row == 3
    # Freeze panes locks the header row in place
    assert ws.freeze_panes == "A2"
    # Auto-filter spans the whole table
    assert ws.auto_filter.ref == ws.dimensions


def test_write_merged_excel_preserves_existing_sira(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """When a record already carries a Sıra (e.g. after distribution),
    the writer keeps it instead of renumbering — the per-team Excel
    needs to point back at the original merged-table row number."""
    monkeypatch.setattr(core, "TARGET_SCHEMA", ["Müşteri"])

    out = tmp_path / "team.xlsx"
    write_merged_excel(
        records=[{"Sıra": 42, "Müşteri": "Ali"}, {"Sıra": 99, "Müşteri": "Veli"}],
        source_files=["x.pdf", "y.pdf"],
        out_path=out,
    )
    wb = load_workbook(out)
    ws = wb.active
    sira_col = [ws.cell(row=r, column=1).value for r in (2, 3)]
    assert sira_col == [42, 99]
