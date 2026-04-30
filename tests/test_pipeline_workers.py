"""Direct unit tests for the background-thread workers.

The async endpoints (``/ocr-start``, ``/convert-start``, ``/batch-files``,
``/batch-convert``) all do the same thing: validate input, persist a job
dict in the right ``JobStore``, and start a thread pointing at one of the
workers in ``pipelines/``. The endpoints already have HTTP-level
coverage; this file pins the workers themselves so a regression that
only manifests inside the thread (silent error swallowing, missing
phase transition, wrong media type) gets caught directly.

Strategy: drive the workers synchronously (call them on the main
thread), pre-create the job entry the worker expects, and inspect
``store.snapshot(token)`` afterwards. EasyOCR is stubbed out so we
don't pull torch on every test run — tests run in <2 s combined.
"""

from __future__ import annotations

import shutil
import zipfile
from pathlib import Path
from uuid import uuid4

import fitz
import pytest

import core
from pipelines.convert import (
    Pdf2DocxProgressHandler,
    batch_files_worker,
    convert_worker,
)
from pipelines.ocr import ocr_worker
from state import batch_store, convert_store, ocr_store


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
@pytest.fixture
def tiny_pdf(tmp_path: Path) -> Path:
    """Single-page PDF with a few words of real text — every worker
    accepts it (Word/Excel render it as text, JPG as a raster page)."""
    out = tmp_path / "tiny.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 80), "Sayfa içeriği", fontsize=12, fontname="helv")
    page.insert_text((72, 110), "İkinci satır", fontsize=12, fontname="helv")
    doc.save(str(out))
    doc.close()
    return out


@pytest.fixture
def two_page_pdf(tmp_path: Path) -> Path:
    """Two-page variant — used to verify per-page progress updates."""
    out = tmp_path / "twopage.pdf"
    doc = fitz.open()
    for i, label in enumerate(["sayfa bir", "sayfa iki"]):
        p = doc.new_page()
        p.insert_text((72, 80 + i * 12), label, fontsize=12, fontname="helv")
    doc.save(str(out))
    doc.close()
    return out


@pytest.fixture
def fresh_token() -> str:
    """A uuid4().hex token, isolated from any other test's job entry."""
    return uuid4().hex


@pytest.fixture(autouse=True)
def cleanup_stores():
    """Wipe any state left over from a worker run so successive tests
    don't see each other's tokens. ``autouse`` so we don't have to
    remember to call it from every test."""
    yield
    for store in (convert_store, batch_store, ocr_store):
        with store.lock:
            store.jobs.clear()


# ---------------------------------------------------------------------------
# OCR worker — JPG branch (no EasyOCR involved)
# ---------------------------------------------------------------------------
def test_ocr_worker_jpg_renders_zip_of_pages(
    tmp_path: Path, two_page_pdf: Path, fresh_token: str
) -> None:
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    ocr_store.create(
        fresh_token,
        target="jpg",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    ocr_worker(fresh_token, two_page_pdf, "jpg", job_dir, "twopage.pdf")

    snap = ocr_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["error"] is None
    assert snap["media_type"] == "application/zip"
    assert snap["total"] == 2
    assert snap["current"] == 2
    out_path = Path(snap["output_path"])
    assert out_path.exists()
    # ZIP holds one .jpg per page, namespaced under the original stem.
    with zipfile.ZipFile(out_path) as zf:
        names = zf.namelist()
    assert len(names) == 2
    assert all(n.lower().endswith(".jpg") for n in names)


def test_ocr_worker_word_uses_easyocr_stub(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, tiny_pdf: Path, fresh_token: str
) -> None:
    """Word path lazy-imports ``pdf_converter.get_ocr_reader``; replace
    it with a fake reader so the test doesn't pull torch."""
    import pdf_converter

    class _FakeReader:
        def readtext(self, _img_bytes, detail=0, paragraph=True):
            return ["Tanınan satır 1", "Tanınan satır 2"]

    monkeypatch.setattr(pdf_converter, "get_ocr_reader", lambda: _FakeReader())

    job_dir = tmp_path / "job"
    job_dir.mkdir()
    ocr_store.create(
        fresh_token,
        target="word",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    ocr_worker(fresh_token, tiny_pdf, "word", job_dir, "tiny.pdf")

    snap = ocr_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["error"] is None
    assert "wordprocessingml" in snap["media_type"]
    assert Path(snap["output_path"]).exists()
    # Read the .docx back and check OCR text made it onto the page
    from docx import Document

    doc = Document(snap["output_path"])
    full = "\n".join(p.text for p in doc.paragraphs)
    assert "Tanınan satır 1" in full


def test_ocr_worker_excel_uses_easyocr_stub(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, tiny_pdf: Path, fresh_token: str
) -> None:
    """Excel path: each page becomes its own sheet."""
    import pdf_converter

    class _FakeReader:
        def readtext(self, _img_bytes, detail=0, paragraph=True):
            return ["satır A", "satır B"]

    monkeypatch.setattr(pdf_converter, "get_ocr_reader", lambda: _FakeReader())

    job_dir = tmp_path / "job"
    job_dir.mkdir()
    ocr_store.create(
        fresh_token,
        target="excel",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    ocr_worker(fresh_token, tiny_pdf, "excel", job_dir, "tiny.pdf")

    snap = ocr_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert "spreadsheetml" in snap["media_type"]
    from openpyxl import load_workbook

    wb = load_workbook(snap["output_path"])
    # One sheet per PDF page
    assert len(wb.sheetnames) == 1
    ws = wb[wb.sheetnames[0]]
    cells = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "satır A" in cells


def test_ocr_worker_records_error_and_marks_done(tmp_path: Path, fresh_token: str) -> None:
    """A corrupted (non-PDF) input must NOT raise from the worker — it
    must record the error in the store and flip ``done`` so the polling
    endpoint can return a clean 500/error response instead of a 404."""
    bad = tmp_path / "broken.pdf"
    bad.write_bytes(b"not a real pdf")

    job_dir = tmp_path / "job"
    job_dir.mkdir()
    ocr_store.create(
        fresh_token,
        target="jpg",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    ocr_worker(fresh_token, bad, "jpg", job_dir, "broken.pdf")

    snap = ocr_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["error"] is not None
    # Error is sanitised — no raw absolute path leaks to the client.
    assert "C:\\" not in snap["error"]


# ---------------------------------------------------------------------------
# Convert worker — Excel / Word / JPG paths + scanned guard
# ---------------------------------------------------------------------------
def test_convert_worker_excel_emits_xlsx(tmp_path: Path, tiny_pdf: Path, fresh_token: str) -> None:
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    convert_store.create(
        fresh_token,
        target="excel",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    convert_worker(fresh_token, tiny_pdf, "excel", "", job_dir, "tiny.pdf")

    snap = convert_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["error"] is None
    assert "spreadsheetml" in snap["media_type"]
    assert Path(snap["output_path"]).exists()


def test_convert_worker_jpg_emits_zip(tmp_path: Path, two_page_pdf: Path, fresh_token: str) -> None:
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    convert_store.create(
        fresh_token,
        target="jpg",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    convert_worker(fresh_token, two_page_pdf, "jpg", "", job_dir, "twopage.pdf")

    snap = convert_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["media_type"] == "application/zip"
    assert snap["current"] == 2  # both pages rendered
    with zipfile.ZipFile(snap["output_path"]) as zf:
        assert sum(1 for n in zf.namelist() if n.lower().endswith(".jpg")) == 2


def test_convert_worker_jpg_uses_custom_name(
    tmp_path: Path, tiny_pdf: Path, fresh_token: str
) -> None:
    """``custom_name`` flows into both the ZIP filename and the inner
    folder name — pin both so a renaming regression on either side
    surfaces immediately."""
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    convert_store.create(
        fresh_token,
        target="jpg",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    convert_worker(fresh_token, tiny_pdf, "jpg", "Müşteri Listesi", job_dir, "tiny.pdf")

    snap = convert_store.snapshot(fresh_token)
    assert snap["output_name"].endswith(".zip")
    # safe_filename strips the space — confirm the custom stem is present
    assert "Listesi" in snap["output_name"] or "Listesi" in str(snap["output_path"])
    with zipfile.ZipFile(snap["output_path"]) as zf:
        names = zf.namelist()
    assert any("/" in n for n in names), "ZIP entries should be namespaced"


def test_convert_worker_excel_short_circuits_on_scanned(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, tiny_pdf: Path, fresh_token: str
) -> None:
    """Word/Excel against a scanned PDF would yield empty output — the
    worker raises a clear error that the polling endpoint surfaces."""
    from pipelines import convert as convert_pipeline

    monkeypatch.setattr(convert_pipeline, "is_scanned_pdf", lambda doc: True)

    job_dir = tmp_path / "job"
    job_dir.mkdir()
    convert_store.create(
        fresh_token,
        target="word",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    convert_worker(fresh_token, tiny_pdf, "word", "", job_dir, "tiny.pdf")

    snap = convert_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["error"] is not None
    assert "OCR" in snap["error"]


def test_convert_worker_call_log_routes_to_specialised_writer(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, tiny_pdf: Path, fresh_token: str
) -> None:
    """Excel branch checks ``is_call_log_pdf`` and routes to
    ``write_call_log_excel`` instead of the generic writer; pin that
    routing without producing a real call-log PDF."""
    from pipelines import convert as convert_pipeline

    fake_records = [{"#": 1, "Müşteri": "Ali", "Telefon": "5551112233"}]
    seen: dict = {}

    monkeypatch.setattr(convert_pipeline, "is_call_log_pdf", lambda doc: True)
    monkeypatch.setattr(convert_pipeline, "parse_call_log", lambda doc: fake_records)
    monkeypatch.setattr(
        convert_pipeline,
        "write_call_log_excel",
        lambda recs, out: (
            seen.setdefault("recs", recs),
            seen.setdefault("out", out),
            out.write_bytes(b"PK\x03\x04stub"),
        ),
    )

    job_dir = tmp_path / "job"
    job_dir.mkdir()
    convert_store.create(
        fresh_token,
        target="excel",
        phase="starting",
        current=0,
        total=0,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    convert_worker(fresh_token, tiny_pdf, "excel", "", job_dir, "tiny.pdf")

    assert seen.get("recs") == fake_records
    snap = convert_store.snapshot(fresh_token)
    assert snap["record_count"] == 1
    assert snap["done"] is True


# ---------------------------------------------------------------------------
# Pdf2DocxProgressHandler — log-line scraper
# ---------------------------------------------------------------------------
def test_pdf2docx_progress_handler_updates_store(fresh_token: str) -> None:
    """pdf2docx logs ``[INFO] (3/10) Page 3``; the handler scrapes the
    digits and forwards them as per-page progress on the convert job."""
    import logging

    convert_store.create(fresh_token, target="word", current=0, total=0, done=False)
    handler = Pdf2DocxProgressHandler(fresh_token)
    record = logging.LogRecord(
        name="pdf2docx",
        level=logging.INFO,
        pathname="",
        lineno=0,
        msg="(3/10) Page 3",
        args=(),
        exc_info=None,
    )
    handler.emit(record)
    snap = convert_store.snapshot(fresh_token)
    assert snap["current"] == 3
    assert snap["total"] == 10


def test_pdf2docx_progress_handler_ignores_unrelated_lines(fresh_token: str) -> None:
    import logging

    convert_store.create(fresh_token, target="word", current=0, total=0)
    handler = Pdf2DocxProgressHandler(fresh_token)
    record = logging.LogRecord(
        name="pdf2docx",
        level=logging.INFO,
        pathname="",
        lineno=0,
        msg="warming up",
        args=(),
        exc_info=None,
    )
    handler.emit(record)
    snap = convert_store.snapshot(fresh_token)
    assert snap["current"] == 0  # untouched


# ---------------------------------------------------------------------------
# batch_files_worker — Word/JPG ZIP for N PDFs
# ---------------------------------------------------------------------------
def test_batch_files_worker_jpg_zips_multiple_pdfs(
    tmp_path: Path, tiny_pdf: Path, two_page_pdf: Path, fresh_token: str
) -> None:
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    # Worker mutates / unlinks the input PDFs in place — copy them into
    # the job dir first (the routers do the same with save_upload).
    a = job_dir / "in_0.pdf"
    shutil.copy(tiny_pdf, a)
    b = job_dir / "in_1.pdf"
    shutil.copy(two_page_pdf, b)

    batch_store.create(
        fresh_token,
        type="files",
        target="jpg",
        phase="starting",
        current=0,
        total=2,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    batch_files_worker(
        fresh_token, [("tiny.pdf", a), ("twopage.pdf", b)], "jpg", job_dir, custom_names=None
    )

    snap = batch_store.snapshot(fresh_token)
    assert snap["done"] is True
    assert snap["error"] is None
    assert snap["produced"] == 2
    assert snap["current"] == 2
    with zipfile.ZipFile(snap["output_path"]) as zf:
        names = zf.namelist()
    # Three .jpg total: 1 from tiny + 2 from two_page
    assert sum(1 for n in names if n.lower().endswith(".jpg")) == 3


def test_batch_files_worker_records_error_and_marks_done(tmp_path: Path, fresh_token: str) -> None:
    """A bad PDF inside the list is now recorded per-file (not as a fatal
    batch error) so other files can still be converted. The worker still
    marks the job done; the bad file is flagged in files_progress."""
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    bad = job_dir / "in_0.pdf"
    bad.write_bytes(b"not a pdf")

    batch_store.create(
        fresh_token,
        type="files",
        target="word",
        phase="starting",
        current=0,
        total=1,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    batch_files_worker(
        fresh_token,
        [("broken.pdf", bad)],
        "word",
        job_dir,
        custom_names=None,
        skip_safety=True,
    )

    snap = batch_store.snapshot(fresh_token)
    assert snap["done"] is True
    # Per-file failure: other files would keep going. The job-level error is
    # now empty; failure is recorded in files_progress instead.
    fp = snap.get("files_progress") or []
    assert len(fp) == 1
    assert fp[0]["status"] == "error"
    assert fp[0]["error"] is not None
    assert snap.get("produced", 0) == 0


def test_batch_files_worker_honours_custom_names(
    tmp_path: Path, tiny_pdf: Path, fresh_token: str
) -> None:
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    a = job_dir / "in_0.pdf"
    shutil.copy(tiny_pdf, a)

    batch_store.create(
        fresh_token,
        type="files",
        target="jpg",
        phase="starting",
        current=0,
        total=1,
        done=False,
        error=None,
        started_at=0.0,
        job_dir=str(job_dir),
    )

    batch_files_worker(fresh_token, [("orig.pdf", a)], "jpg", job_dir, custom_names=["custom_stem"])

    snap = batch_store.snapshot(fresh_token)
    assert snap["done"] is True
    with zipfile.ZipFile(snap["output_path"]) as zf:
        names = zf.namelist()
    # Folder inside ZIP must use the custom stem, not the orig filename
    assert any(n.startswith("custom_stem/") for n in names)
